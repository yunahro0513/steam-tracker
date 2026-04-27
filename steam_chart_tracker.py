#!/usr/bin/env python3
"""
Steam 인기 차트 자동 추적기
매일 실행하여 상위 50개 게임의 지표를 Excel에 누적 저장합니다.

추적 지표:
  - 동시 접속자 수 (CCU)
  - 긍정 리뷰 비율 (%)
  - 현재 가격 및 할인율

롱런 기준:
  - 2주(14일) 이상 상위 50위 유지
  - 4주(28일) 이상 상위 50위 유지
"""

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date
import time
import os
import json

# ── 설정 ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH   = os.path.join(SCRIPT_DIR, "steam_chart_tracker.xlsx")
JSON_PATH    = os.path.join(SCRIPT_DIR, "docs", "data.json")
TOP_N        = 50   # 상위 50개 게임 추적
LONGRUN_2W   = 14   # 2주(14일) 이상
LONGRUN_4W   = 28   # 4주(28일) 이상


# ── 데이터 수집 ───────────────────────────────────────────────────────────────

def fetch_steamspy_top100():
    """SteamSpy API: 최근 2주 상위 100개 게임"""
    url = "https://steamspy.com/api.php?request=top100in2weeks"
    r = requests.get(url, timeout=30)
    r.raise_for_status()
    data = r.json()

    games = []
    for rank, (appid, info) in enumerate(data.items(), 1):
        pos = info.get("positive", 0) or 0
        neg = info.get("negative", 0) or 0
        total = pos + neg
        review_pct = round(pos / total * 100, 1) if total > 0 else 0

        games.append({
            "rank":             rank,
            "appid":            int(appid),
            "name":             info.get("name", ""),
            "ccu":              info.get("ccu", 0) or 0,
            "review_score_pct": review_pct,
            "total_reviews":    total,
            "positive_reviews": pos,
        })

    return games[:TOP_N]


def fetch_steam_price(appid):
    """Steam Store API: 가격 및 할인율"""
    url = (
        f"https://store.steampowered.com/api/appdetails/"
        f"?appids={appid}&cc=kr&filters=price_overview"
    )
    try:
        r = requests.get(url, timeout=10)
        d = r.json()
        app = d.get(str(appid), {})
        if app.get("success") and app.get("data"):
            p = app["data"].get("price_overview", {})
            return {
                "price_krw":          p.get("final", 0) // 100,
                "discount_pct":       p.get("discount_percent", 0),
                "original_price_krw": p.get("initial", 0) // 100,
            }
    except Exception:
        pass
    return {"price_krw": None, "discount_pct": 0, "original_price_krw": None}


def collect_today_data():
    """오늘의 차트 전체 수집"""
    print("▶ SteamSpy 상위 50 게임 수집 중...")
    games = fetch_steamspy_top100()

    today_str = date.today().isoformat()
    rows = []
    for i, g in enumerate(games, 1):
        print(f"  [{i:2d}/{len(games)}] {g['name'][:40]}")
        price = fetch_steam_price(g["appid"])
        time.sleep(0.4)

        rows.append({
            "date":               today_str,
            "rank":               g["rank"],
            "appid":              g["appid"],
            "name":               g["name"],
            "ccu":                g["ccu"],
            "review_score_pct":   g["review_score_pct"],
            "total_reviews":      g["total_reviews"],
            "positive_reviews":   g["positive_reviews"],
            "price_krw":          price["price_krw"],
            "discount_pct":       price["discount_pct"],
            "original_price_krw": price["original_price_krw"],
        })

    return pd.DataFrame(rows)


# ── 분석 ──────────────────────────────────────────────────────────────────────

def analyze_longrun(df, min_days):
    """min_days 이상 상위 50위를 유지한 게임 집계"""
    if df.empty:
        return pd.DataFrame()

    df = df.copy()
    df["date"] = pd.to_datetime(df["date"])

    stats = df.groupby(["appid", "name"]).agg(
        days_in_top         = ("date",             "nunique"),
        avg_rank            = ("rank",             "mean"),
        best_rank           = ("rank",             "min"),
        avg_ccu             = ("ccu",              "mean"),
        latest_ccu          = ("ccu",              "last"),
        avg_review_score    = ("review_score_pct", "mean"),
        latest_review_score = ("review_score_pct", "last"),
        total_reviews       = ("total_reviews",    "last"),
        latest_price        = ("price_krw",        "last"),
        max_discount        = ("discount_pct",     "max"),
        first_seen          = ("date",             "min"),
        last_seen           = ("date",             "max"),
    ).reset_index()

    result = stats[stats["days_in_top"] >= min_days].copy()
    result.sort_values("days_in_top", ascending=False, inplace=True)

    result["avg_rank"]         = result["avg_rank"].round(1)
    result["avg_ccu"]          = result["avg_ccu"].round(0).astype(int)
    result["avg_review_score"] = result["avg_review_score"].round(1)
    result["first_seen"]       = result["first_seen"].dt.strftime("%Y-%m-%d")
    result["last_seen"]        = result["last_seen"].dt.strftime("%Y-%m-%d")

    return result


# ── Excel 출력 ────────────────────────────────────────────────────────────────

HDR_FILL  = PatternFill("solid", start_color="1F4E79")  # 진한 남색
HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL  = PatternFill("solid", start_color="EBF3FB")  # 연한 파랑
DIS_FILL  = PatternFill("solid", start_color="C6EFCE")  # 연한 초록 (할인)
LRN2_FILL = PatternFill("solid", start_color="FFF3CD")  # 연한 노랑 (2주+)
LRN4_FILL = PatternFill("solid", start_color="FFD700")  # 골드 (4주+)
CENTER    = Alignment(horizontal="center", vertical="center")

L_COLS = {
    "name":                "게임명",
    "days_in_top":         "유지 일수",
    "avg_rank":            "평균 순위",
    "best_rank":           "최고 순위",
    "avg_ccu":             "평균 동접",
    "latest_ccu":          "최근 동접",
    "avg_review_score":    "평균 긍정리뷰(%)",
    "latest_review_score": "최근 긍정리뷰(%)",
    "total_reviews":       "총 리뷰수",
    "latest_price":        "현재 가격(₩)",
    "max_discount":        "최대 할인율(%)",
    "first_seen":          "첫 관측일",
    "last_seen":           "최근 관측일",
}


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_longrun_sheet(ws, title, longrun_df, row_fill, min_days):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13, color="B8860B")
    ws.append([])

    ws.append(list(L_COLS.values()))
    _style_header(ws, 3, len(L_COLS))

    if not longrun_df.empty:
        for ri, row in enumerate(longrun_df.itertuples(index=False), 4):
            for ci, k in enumerate(L_COLS.keys(), 1):
                ws.cell(ri, ci, value=getattr(row, k, None))
            for ci in range(1, len(L_COLS) + 1):
                ws.cell(ri, ci).fill = row_fill
    else:
        ws["A4"] = (
            f"아직 {min_days}일 분량의 데이터가 쌓이지 않았습니다. "
            "매일 자동 수집이 진행되면 이 시트가 채워집니다."
        )
        ws["A4"].font = Font(italic=True, color="888888")

    _set_col_widths(ws, [36, 10, 10, 10, 12, 12, 16, 16, 12, 14, 14, 14, 14])
    ws.freeze_panes = "A4"


def build_excel(all_df, today_df, longrun_2w_df, longrun_4w_df):
    wb = Workbook()

    # ── 시트 1: 일별 스냅샷 ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "일별 스냅샷"

    COLS = {
        "date":               "날짜",
        "rank":               "순위",
        "appid":              "AppID",
        "name":               "게임명",
        "ccu":                "동시접속자",
        "review_score_pct":   "긍정리뷰(%)",
        "total_reviews":      "총 리뷰수",
        "positive_reviews":   "긍정리뷰수",
        "price_krw":          "가격(₩)",
        "discount_pct":       "할인율(%)",
        "original_price_krw": "정가(₩)",
    }
    ws1.append(list(COLS.values()))
    _style_header(ws1, 1, len(COLS))

    keys = list(COLS.keys())
    for ri, row in enumerate(all_df.itertuples(index=False), 2):
        for ci, k in enumerate(keys, 1):
            ws1.cell(ri, ci, value=getattr(row, k, None))
        if ri % 2 == 0:
            for ci in range(1, len(keys) + 1):
                ws1.cell(ri, ci).fill = ALT_FILL

    _set_col_widths(ws1, [12, 5, 10, 36, 12, 12, 12, 12, 10, 10, 10])
    ws1.freeze_panes = "A2"

    # ── 시트 2: 오늘의 차트 ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("오늘의 차트")
    ws2["A1"] = f"Steam 인기 차트 — {date.today().isoformat()}"
    ws2["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws2.append([])

    T_COLS = ["순위", "게임명", "동시접속자", "긍정리뷰(%)", "총 리뷰수", "가격(₩)", "할인율(%)"]
    T_KEYS = ["rank", "name", "ccu", "review_score_pct", "total_reviews", "price_krw", "discount_pct"]
    ws2.append(T_COLS)
    _style_header(ws2, 3, len(T_COLS))

    for ri, row in enumerate(today_df.itertuples(index=False), 4):
        for ci, k in enumerate(T_KEYS, 1):
            ws2.cell(ri, ci, value=getattr(row, k, None))
        disc = getattr(row, "discount_pct", 0) or 0
        base_fill = DIS_FILL if disc > 0 else (ALT_FILL if ri % 2 == 0 else None)
        if base_fill:
            for ci in range(1, len(T_KEYS) + 1):
                ws2.cell(ri, ci).fill = base_fill

    _set_col_widths(ws2, [5, 36, 12, 12, 12, 10, 10])
    ws2.freeze_panes = "A4"

    # ── 시트 3: 2주+ 롱런 분석 ──────────────────────────────────────────────
    ws3 = wb.create_sheet("2주+ 롱런 분석")
    _write_longrun_sheet(
        ws3,
        f"2주(14일)+ 상위 50위 유지 게임 — 기준일: {date.today().isoformat()}",
        longrun_2w_df,
        LRN2_FILL,
        LONGRUN_2W,
    )

    # ── 시트 4: 4주+ 롱런 분석 ──────────────────────────────────────────────
    ws4 = wb.create_sheet("4주+ 롱런 분석")
    _write_longrun_sheet(
        ws4,
        f"4주(28일)+ 상위 50위 유지 게임 — 기준일: {date.today().isoformat()}",
        longrun_4w_df,
        LRN4_FILL,
        LONGRUN_4W,
    )

    wb.save(EXCEL_PATH)
    print(f"✔ Excel 저장: {EXCEL_PATH}")


# ── JSON 출력 (GitHub Pages용) ───────────────────────────────────────────────

def write_json(today_df, longrun_2w_df, longrun_4w_df):
    def to_records(df):
        if df.empty:
            return []
        # pandas to_json이 NaN → null 변환을 올바르게 처리함
        return json.loads(df.to_json(orient="records", force_ascii=False))

    data = {
        "updated":    date.today().isoformat(),
        "today_chart": to_records(today_df[["rank","appid","name","ccu","review_score_pct","total_reviews","price_krw","discount_pct"]]),
        "longrun_2w": to_records(longrun_2w_df),
        "longrun_4w": to_records(longrun_4w_df),
    }
    os.makedirs(os.path.dirname(JSON_PATH), exist_ok=True)
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"✔ JSON 저장: {JSON_PATH}")


# ── 메인 ──────────────────────────────────────────────────────────────────────

def load_existing(path):
    if not os.path.exists(path):
        return pd.DataFrame()
    try:
        df = pd.read_excel(path, sheet_name="일별 스냅샷")
        df["date"] = pd.to_datetime(df["date"]).dt.date.astype(str)
        return df
    except Exception as e:
        print(f"기존 파일 로드 실패 ({e}), 새로 시작합니다.")
        return pd.DataFrame()


def main():
    today_str = date.today().isoformat()
    print(f"\n{'='*55}")
    print(f"  Steam 차트 추적기  |  {today_str}")
    print(f"{'='*55}")

    today_df = collect_today_data()
    existing = load_existing(EXCEL_PATH)

    if not existing.empty:
        existing = existing[existing["date"] != today_str]
    all_df = pd.concat([existing, today_df], ignore_index=True) if not existing.empty else today_df

    longrun_2w = analyze_longrun(all_df.copy(), LONGRUN_2W)
    longrun_4w = analyze_longrun(all_df.copy(), LONGRUN_4W)
    print(f"\n▶ 2주+ 롱런 게임: {len(longrun_2w)}개")
    print(f"▶ 4주+ 롱런 게임: {len(longrun_4w)}개")

    build_excel(all_df, today_df, longrun_2w, longrun_4w)
    write_json(today_df, longrun_2w, longrun_4w)
    print("  완료!\n")


if __name__ == "__main__":
    main()
